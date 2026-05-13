"""
STEP 10: NEW NON-TRADES BATCH — 10 Fresh Leads

One-off script that fetches 10 new leads from underrepresented sectors
not well-covered in the main pipeline run (IT, accounting, manufacturing,
veterinary, optometry, physiotherapy, financial planning, printing, etc.).

Explicitly EXCLUDES:
  - All trades: plumbing, electrical, roofing, HVAC
  - Any place_id already in data/raw_candidates.csv

Output: data/Oakville_NewBatch_NonTrades_10.xlsx
"""

import os
import sys
import json
import re
import math
import time
import hashlib
import requests
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(__file__))
from config import (
    GOOGLE_PLACES_API_KEY,
    GEO_BOUNDS,
    GOOGLE_DELAY_SECONDS,
    TARGET_CITY,
    TARGET_PROVINCE,
    TARGET_POSTAL_PREFIXES,
    EXCLUDED_SECTOR_KEYWORDS,
    HIGH_VALUE_SECTOR_KEYWORDS,
    WEIGHT_YEARS_IN_BUSINESS,
    WEIGHT_REVIEW_COUNT,
    WEIGHT_SECTOR_SIGNAL,
    WEIGHT_EMPLOYEE_COUNT,
    WEIGHT_DATA_QUALITY,
    WEIGHT_WEBSITE_PRESENCE,
    WEIGHT_LOCATION_BONUS,
    WEIGHT_SDE_FIT_SIGNAL,
    TARGET_SDE_LOW,
    TARGET_SDE_HIGH,
    REVIEW_THRESHOLDS,
    REVENUE_ESTIMATION,
    INDUSTRY_MARGINS,
    INDUSTRY_REVENUE_PER_EMPLOYEE,
    CHAIN_FILTER_ENABLED,
    CACHE_ENABLED,
    CHECKPOINT_RAW,
    QUALIFICATION_THRESHOLD,
)
from utils.chain_filter import is_chain_or_franchise
from utils.subsidiary_detector import is_subsidiary
from utils.employee_estimator import estimate_employee_range, estimate_business_age

TEXT_URL = "https://maps.googleapis.com/maps/api/place/textsearch/json"
CACHE_DIR = "data/cache/acquisition"
OUTPUT_EXCEL = "data/Oakville_NewBatch_NonTrades_10.xlsx"

# ── New keywords targeting underrepresented non-trades sectors ────────────────
# These are different from ACQUISITION_KEYWORDS in config.py to find fresh leads.
NEW_BATCH_KEYWORDS = [
    # IT / technology
    "IT services company",
    "managed services provider",
    "software company",
    "technology consulting",
    # Accounting / financial
    "accounting firm",
    "CPA firm",
    "financial planning firm",
    "wealth management",
    "insurance brokerage",
    # Manufacturing
    "manufacturing company",
    "food manufacturing",
    "packaging company",
    "precision machining",
    # Healthcare (non-trades)
    "veterinary clinic",
    "optometry clinic",
    "physiotherapy clinic",
    "chiropractic clinic",
    "medical clinic",
    # Specialized services
    "environmental consulting",
    "marketing agency",
    "printing company",
    "industrial supply",
    "laboratory services",
    "staffing company",
]

# Trades terms to exclude (in addition to config exclusions)
TRADES_EXCLUSION_TERMS = [
    "plumbing", "plumber", "electrical", "electrician",
    "roofing", "roofer", "hvac", "heating", "cooling",
    "air conditioning", "ventilation",
]


# ── Cache helpers ─────────────────────────────────────────────────────────────

def cache_key(params_str):
    h = hashlib.md5(params_str.encode()).hexdigest()[:12]
    return os.path.join(CACHE_DIR, f"nb10_{h}.json")


def cached_request(url, params):
    os.makedirs(CACHE_DIR, exist_ok=True)
    key_str = json.dumps(params, sort_keys=True)
    c_path = cache_key(key_str)

    if CACHE_ENABLED and os.path.exists(c_path):
        with open(c_path, "r") as f:
            return json.load(f), True

    try:
        resp = requests.get(url, params=params, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        if CACHE_ENABLED and data.get("status") in ("OK", "ZERO_RESULTS"):
            with open(c_path, "w") as f:
                json.dump(data, f)
        return data, False
    except requests.RequestException as e:
        print(f"    [ERROR] {e}")
        return None, False


def fetch_keyword(keyword, existing_ids):
    """Fetch up to 3 pages for a single keyword, return new place records."""
    query = f"{keyword} {TARGET_CITY} {TARGET_PROVINCE}"
    params = {"query": query, "key": GOOGLE_PLACES_API_KEY}

    new_records = []
    data, was_cached = cached_request(TEXT_URL, params)
    if not data or data.get("status") not in ("OK", "ZERO_RESULTS"):
        return new_records, not was_cached

    results = data.get("results", [])
    page = 1
    while "next_page_token" in data and page < 3:
        page += 1
        time.sleep(2.0)
        next_params = {"pagetoken": data["next_page_token"], "key": GOOGLE_PLACES_API_KEY}
        data, _ = cached_request(TEXT_URL, next_params)
        if data and data.get("status") == "OK":
            results.extend(data.get("results", []))
        else:
            break

    for r in results:
        pid = r.get("place_id")
        if pid and pid not in existing_ids:
            loc = r.get("geometry", {}).get("location", {})
            new_records.append({
                "google_place_id": pid,
                "company_name": r.get("name", ""),
                "address_raw": r.get("formatted_address", r.get("vicinity", "")),
                "lat": loc.get("lat"),
                "lng": loc.get("lng"),
                "google_types": ",".join(r.get("types", [])),
                "google_rating": r.get("rating"),
                "review_count": r.get("user_ratings_total", 0),
                "business_status": r.get("business_status", ""),
                "price_level": r.get("price_level"),
                "num_employees": None,
                "website": None,
                "phone": None,
                "owner_name": None,
                "owner_confidence": None,
                "owner_source": None,
                "registration_date": None,
                "estimated_years": None,
            })
            existing_ids.add(pid)

    return new_records, not was_cached


# ── Filtering ─────────────────────────────────────────────────────────────────

def is_in_oakville(row):
    addr = str(row.get("address_raw", "")).upper()
    lat = row.get("lat")
    lng = row.get("lng")
    has_postal = any(p in addr for p in TARGET_POSTAL_PREFIXES)
    in_bounds = (
        lat is not None and lng is not None and
        GEO_BOUNDS["south"] <= lat <= GEO_BOUNDS["north"] and
        GEO_BOUNDS["west"] <= lng <= GEO_BOUNDS["east"]
    )
    return has_postal or in_bounds


def is_excluded_sector(row):
    text = (str(row.get("company_name", "")) + " " +
            str(row.get("google_types", ""))).lower()
    for kw in EXCLUDED_SECTOR_KEYWORDS:
        if kw.lower() in text:
            return True
    return False


def is_trades(row):
    """Return True if the business looks like a trades company."""
    text = (str(row.get("company_name", "")) + " " +
            str(row.get("google_types", ""))).lower()
    for term in TRADES_EXCLUSION_TERMS:
        if term.lower() in text:
            return True
    return False


def apply_filters(df):
    before = len(df)

    # Geography
    df = df[df.apply(is_in_oakville, axis=1)].copy()
    print(f"    After geo filter: {len(df)} (removed {before - len(df)})")

    # Permanently closed
    df = df[df["business_status"] != "CLOSED_PERMANENTLY"].copy()

    # Sector exclusions (from config)
    df = df[~df.apply(is_excluded_sector, axis=1)].copy()
    print(f"    After sector exclusion: {len(df)}")

    # Trades exclusion (user request)
    before_trades = len(df)
    df = df[~df.apply(is_trades, axis=1)].copy()
    print(f"    After trades exclusion: {len(df)} (removed {before_trades - len(df)})")

    # Chain/franchise filter
    if CHAIN_FILTER_ENABLED:
        before_chain = len(df)
        mask = df.apply(lambda r: is_chain_or_franchise(
            str(r.get("company_name", "")),
            str(r.get("google_types", ""))
        )[0], axis=1)
        df = df[~mask].copy()
        print(f"    After chain filter: {len(df)} (removed {before_chain - len(df)})")

    # Subsidiary filter
    before_sub = len(df)
    mask = df.apply(lambda r: is_subsidiary(
        str(r.get("company_name", "")),
        str(r.get("google_types", ""))
    )[0], axis=1)
    df = df[~mask].copy()
    print(f"    After subsidiary filter: {len(df)} (removed {before_sub - len(df)})")

    # Filter individual agents (single-person names without a business suffix)
    # These are not acquisition targets — they lack employees and revenue scale.
    _INDIVIDUAL_INDICATORS = [
        "insurance agent", "desjardins agent", "financial advisor",
        "investment advisor", "re/max", "royal lepage", "century 21",
    ]
    _CORP_SUFFIXES = re.compile(
        r'\b(inc|ltd|corp|group|clinic|centre|center|services|solutions|'
        r'consulting|management|systems|technologies|health|medical|associates)\b',
        re.IGNORECASE
    )
    def is_individual_agent(row):
        name = str(row.get("company_name", "")).lower()
        types = str(row.get("google_types", "")).lower()
        # Flag if name matches individual indicator
        for ind in _INDIVIDUAL_INDICATORS:
            if ind in name:
                return True
        # Flag if no corporate suffix AND looks like a person's name (no common biz words)
        has_corp = bool(_CORP_SUFFIXES.search(name))
        # Count words — 2-word names without corporate suffix are likely individuals
        words = [w for w in name.split() if len(w) > 1]
        if not has_corp and len(words) <= 2 and "insurance" in types:
            return True
        return False

    before_agents = len(df)
    df = df[~df.apply(is_individual_agent, axis=1)].copy()
    print(f"    After individual agent filter: {len(df)} (removed {before_agents - len(df)})")

    # Deduplicate by normalized company name (keep first occurrence = highest scoring later)
    before_dedup = len(df)
    def normalize_name(name):
        name = re.sub(r'\b(inc|ltd|corp|llc|co|group)\b\.?', '', str(name).lower())
        name = re.sub(r'[^a-z0-9 ]', ' ', name)
        return ' '.join(name.split())
    df["_norm_name"] = df["company_name"].apply(normalize_name)
    df = df.drop_duplicates(subset=["_norm_name"]).drop(columns=["_norm_name"])
    print(f"    After name dedup: {len(df)} (removed {before_dedup - len(df)})")

    return df


# ── Scoring (mirrors 06_score.py logic) ──────────────────────────────────────

def get_industry_label(row):
    text = (str(row.get("company_name", "")).lower() + " " +
            str(row.get("google_types", "")).lower())
    industry_map = [
        (["manufactur", "fabricat", "factory", "production", "machining", "cnc"], "manufacturing"),
        (["plumbing", "hvac", "electrical", "roofing", "welding", "heating"], "trades"),
        (["construction", "renovation", "contractor", "builder"], "construction"),
        (["transport", "trucking", "freight", "logistics", "moving"], "transportation"),
        (["wholesale", "distributor", "distribution", "supply"], "wholesale"),
        (["cleaning", "janitorial"], "cleaning_services"),
        (["landscaping", "lawn", "property maintenance"], "landscaping"),
        (["it service", "software", "managed service", "technology"], "it_services"),
        (["engineering", "architect", "consulting"], "professional_services"),
        (["equipment", "rental", "leasing"], "equipment_rental"),
        (["printing", "print shop", "packaging"], "printing"),
        (["food processing", "food manufacturing"], "food_processing"),
        (["dental", "physio", "veterinar", "chiropract", "medical", "optom"], "healthcare_services"),
        (["accounting", "financial", "insurance", "bookkeep", "wealth", "cpa"], "financial_services"),
        (["staffing", "recruitment", "placement"], "staffing"),
    ]
    for keywords, label in industry_map:
        if any(kw in text for kw in keywords):
            return label
    return "general"


def estimate_revenue(row, industry="default"):
    mids = []
    confidence = 20

    emp = row.get("num_employees")
    if emp is not None and pd.notna(emp) and float(emp) > 0:
        emp = float(emp)
        rev_per_emp = INDUSTRY_REVENUE_PER_EMPLOYEE.get(
            industry, INDUSTRY_REVENUE_PER_EMPLOYEE["default"])
        mids.append(emp * rev_per_emp[1])
        confidence += 25

    reviews = row.get("review_count", 0)
    if reviews and pd.notna(reviews) and float(reviews) > 0:
        reviews = float(reviews)
        if reviews >= REVIEW_THRESHOLDS["excellent"]:
            mids.append(3_000_000)
        elif reviews >= REVIEW_THRESHOLDS["good"]:
            mids.append(2_000_000)
        elif reviews >= REVIEW_THRESHOLDS["moderate"]:
            mids.append(1_500_000)
        elif reviews >= REVIEW_THRESHOLDS["low"]:
            mids.append(1_000_000)
        else:
            mids.append(700_000)
        confidence += 10

    age_factor = 1.0
    est_years = row.get("estimated_years")
    if est_years is not None and pd.notna(est_years):
        y = float(est_years)
        if y >= 25:
            age_factor = 1.3
        elif y >= 15:
            age_factor = 1.1
        elif y >= 10:
            age_factor = 1.0
        else:
            age_factor = 0.8
        confidence += 10

    if mids:
        avg_mid = (sum(mids) / len(mids)) * age_factor
    else:
        base = REVENUE_ESTIMATION["base_range"]
        avg_mid = (base[0] + base[1]) / 2

    if confidence >= 60:
        margin = REVENUE_ESTIMATION["confidence_margin_high"]
    elif confidence >= 40:
        margin = REVENUE_ESTIMATION["confidence_margin_moderate"]
    else:
        margin = REVENUE_ESTIMATION["confidence_margin_low"]

    return avg_mid * (1 - margin), avg_mid, avg_mid * (1 + margin), min(confidence, 85)


def score_row(row, industry):
    # Years
    est = row.get("estimated_years")
    s_years = min((float(est) / 30.0) * 100, 100.0) if est and pd.notna(est) else 15.0

    # Reviews
    count = float(row.get("review_count", 0) or 0)
    if count >= REVIEW_THRESHOLDS["excellent"]: s_reviews = 100.0
    elif count >= REVIEW_THRESHOLDS["good"]:    s_reviews = 80.0
    elif count >= REVIEW_THRESHOLDS["moderate"]: s_reviews = 60.0
    elif count >= REVIEW_THRESHOLDS["low"]:     s_reviews = 40.0
    elif count >= REVIEW_THRESHOLDS["very_low"]: s_reviews = 25.0
    else:                                        s_reviews = 10.0

    # Sector signal
    text = (str(row.get("company_name", "")) + " " +
            str(row.get("google_types", ""))).lower()
    matches = sum(1 for kw in HIGH_VALUE_SECTOR_KEYWORDS if kw.lower() in text)
    if matches >= 3:    s_sector = 100.0
    elif matches == 2:  s_sector = 80.0
    elif matches == 1:  s_sector = 55.0
    else:               s_sector = 20.0

    # Employee count
    emp = row.get("num_employees")
    if emp is None or pd.isna(emp):
        s_employees = 30.0
    else:
        emp = float(emp)
        if 15 <= emp <= 50:      s_employees = 100.0
        elif 10 <= emp < 15:     s_employees = 75.0
        elif 50 < emp <= 75:     s_employees = 70.0
        elif 5 <= emp < 10:      s_employees = 40.0
        elif 75 < emp <= 100:    s_employees = 50.0
        elif 100 < emp <= 200:   s_employees = 30.0
        else:                    s_employees = 10.0

    # Data quality
    s_quality = 0.0
    for field, weight in [("company_name", 15), ("address_raw", 10), ("phone", 15),
                           ("website", 15), ("google_rating", 10), ("review_count", 5),
                           ("owner_name", 20), ("num_employees", 10)]:
        val = row.get(field)
        if pd.notna(val) and str(val).strip() not in ("", "nan", "None", "N/A", "Not found"):
            s_quality += weight

    # Website
    website = row.get("website")
    s_website = 60.0 if (website and pd.notna(website) and str(website).strip()) else 0.0

    # Location
    addr = str(row.get("address_raw", "")).upper()
    s_location = 100.0 if any(p in addr for p in TARGET_POSTAL_PREFIXES) else 50.0

    # SDE fit
    rev_low, rev_mid, rev_high, conf = estimate_revenue(row, industry)
    margin = INDUSTRY_MARGINS.get(industry, INDUSTRY_MARGINS["default"])
    sde_low = rev_low * margin
    sde_high = rev_high * margin
    if sde_low <= TARGET_SDE_HIGH and sde_high >= TARGET_SDE_LOW:
        overlap_low = max(sde_low, TARGET_SDE_LOW)
        overlap_high = min(sde_high, TARGET_SDE_HIGH)
        overlap_ratio = (overlap_high - overlap_low) / (TARGET_SDE_HIGH - TARGET_SDE_LOW)
        base_score = 50 + (overlap_ratio * 50)
        s_sde = base_score * (0.5 + 0.5 * conf / 100.0)
    elif sde_high < TARGET_SDE_LOW:
        gap_ratio = (TARGET_SDE_LOW - sde_high) / TARGET_SDE_LOW
        s_sde = max(10.0, 40.0 * (1 - gap_ratio))
    else:
        gap_ratio = (sde_low - TARGET_SDE_HIGH) / TARGET_SDE_HIGH
        s_sde = max(10.0, 35.0 * (1 - gap_ratio))

    total = (
        s_years    * WEIGHT_YEARS_IN_BUSINESS +
        s_reviews  * WEIGHT_REVIEW_COUNT +
        s_sector   * WEIGHT_SECTOR_SIGNAL +
        s_employees * WEIGHT_EMPLOYEE_COUNT +
        s_quality  * WEIGHT_DATA_QUALITY +
        s_website  * WEIGHT_WEBSITE_PRESENCE +
        s_location * WEIGHT_LOCATION_BONUS +
        s_sde      * WEIGHT_SDE_FIT_SIGNAL
    )
    return {
        "industry_label": industry,
        "estimated_revenue_low": rev_low,
        "estimated_revenue_high": rev_high,
        "estimated_sde_low": sde_low,
        "estimated_sde_high": sde_high,
        "revenue_confidence": conf,
        "total_score": round(total, 1),
    }


# ── Export helpers ────────────────────────────────────────────────────────────

CATEGORY_RULES = [
    (["machining", "cnc", "precision", "tool and die"], "Precision Manufacturing"),
    (["plastics", "injection", "mold"], "Plastics Manufacturing"),
    (["metal", "steel", "welding", "fabricat"], "Metal Fabrication"),
    (["millwork", "cabinet", "carpentry"], "Wood / Millwork"),
    (["printing", "print", "label", "packaging"], "Printing & Packaging"),
    (["food processing", "food manufacturing"], "Food Processing"),
    (["manufactur", "factory", "production", "assembly"], "Light Manufacturing"),
    (["it service", "software", "managed service", "technology"], "IT / Technology Services"),
    (["engineering", "engineer"], "Engineering Services"),
    (["consulting"], "Consulting / Professional Services"),
    (["staffing", "recruitment", "placement"], "Staffing / Recruitment"),
    (["wholesale", "distributor", "supply", "supplier"], "Wholesale / Distribution"),
    (["cleaning", "janitorial"], "Cleaning Services"),
    (["landscaping", "lawn", "property maintenance"], "Landscaping"),
    (["security", "guard", "alarm", "surveillance"], "Security Services"),
    (["auto body", "collision", "auto repair"], "Automotive Services"),
    (["waste", "recycling", "disposal"], "Waste Management"),
    (["equipment", "rental", "leasing"], "Equipment Services"),
    (["dental", "dentist"], "Dental Practice"),
    (["physio", "rehabilitation"], "Physiotherapy / Rehab"),
    (["veterinary", "vet clinic", "animal"], "Veterinary Services"),
    (["chiropract"], "Chiropractic"),
    (["optom", "optical"], "Optometry"),
    (["medical clinic", "walk-in", "home care", "home health"], "Healthcare Services"),
    (["accounting", "accountant", "cpa", "bookkeep"], "Accounting / Financial"),
    (["wealth", "financial planning", "financial advisor"], "Financial Planning"),
    (["insurance"], "Insurance Services"),
    (["property management", "real estate"], "Property Management"),
    (["marketing", "advertising", "digital marketing"], "Marketing / Advertising"),
    (["environmental", "remediation"], "Environmental Services"),
    (["laboratory", "testing", "inspection"], "Testing / Lab Services"),
]


def classify_category(text):
    text_lower = text.lower()
    for keywords, category in CATEGORY_RULES:
        if any(re.search(r'\b' + re.escape(kw), text_lower) for kw in keywords):
            return category
    return "General Business"


def fmt_money(val):
    if val >= 1_000_000:
        return f"${val/1_000_000:.1f}M"
    elif val >= 1000:
        return f"${val/1000:.0f}K"
    return f"${val:.0f}"


def format_age_range(row):
    est = row.get("estimated_years")
    if est is not None and pd.notna(est):
        y = float(est)
        if y >= 15:   return "15+ years (est.)"
        elif y >= 10: return "10-15 years (est.)"
        elif y >= 5:  return "5-10 years (est.)"
        elif y >= 2:  return "2-5 years (est.)"
        else:         return "<2 years (est.)"
    return "Unknown"


def format_employee_range(emp):
    if emp is None or pd.isna(emp):
        return "Unknown"
    emp = int(float(emp))
    if emp <= 5:    return "1-5"
    elif emp <= 10: return "5-10"
    elif emp <= 20: return "10-20"
    elif emp <= 35: return "15-35"
    elif emp <= 50: return "25-50"
    elif emp <= 75: return "50-75"
    elif emp <= 100: return "75-100"
    elif emp <= 200: return "100-200"
    return f"{emp}+"


def compose_notes(row, category, scored):
    parts = []
    parts.append(f"Category: {category}")

    sde_low = scored.get("estimated_sde_low", 0)
    sde_high = scored.get("estimated_sde_high", 0)
    if sde_low and sde_high:
        if sde_low <= TARGET_SDE_HIGH and sde_high >= TARGET_SDE_LOW:
            parts.append("SDE estimate overlaps target range")
        elif sde_high < TARGET_SDE_LOW:
            parts.append("REVIEW: SDE estimate below target range")
        else:
            parts.append("REVIEW: SDE estimate above target range")

    conf = scored.get("revenue_confidence", 0)
    if conf >= 60:   parts.append("Estimate confidence: HIGH")
    elif conf >= 40: parts.append("Estimate confidence: MODERATE")
    else:            parts.append("Estimate confidence: LOW (verify manually)")

    score = scored.get("total_score", 0)
    if score >= 70:   parts.append("High acquisition fit; priority follow-up")
    elif score >= 55: parts.append("Good acquisition fit; worth investigating")
    else:             parts.append("Moderate fit; review recommended")

    return " | ".join(parts)


def to_standard_row(row, scored):
    desc_text = str(row.get("company_name", "")) + " " + str(row.get("google_types", ""))
    category = classify_category(desc_text)

    rev_low  = scored.get("estimated_revenue_low", 0)
    rev_high = scored.get("estimated_revenue_high", 0)
    sde_low  = scored.get("estimated_sde_low", 0)
    sde_high = scored.get("estimated_sde_high", 0)

    emp = row.get("num_employees")
    emp_str = format_employee_range(emp) if emp and pd.notna(emp) else (
        (estimate_employee_range(row) or {}).get("employee_range", "Unknown")
    )

    return {
        "business_name":          str(row.get("company_name", "")),
        "address":                str(row.get("address_raw", "")),
        "city":                   "Oakville",
        "postal_code":            "",
        "website":                str(row.get("website", "")) if row.get("website") else "",
        "phone":                  str(row.get("phone", "")) if row.get("phone") else "",
        "owner_name":             "Not found",
        "owner_confidence":       "none",
        "owner_source":           "Owner not found; manual research required",
        "industry":               scored.get("industry_label", "general"),
        "category_standardized":  category,
        "employee_range_estimate": emp_str,
        "revenue_range_estimate": f"{fmt_money(rev_low)}-{fmt_money(rev_high)}" if rev_low else "Unknown",
        "sde_range_estimate":     f"{fmt_money(sde_low)}-{fmt_money(sde_high)}" if sde_low else "Unknown",
        "age_range_estimate":     format_age_range(row),
        "acquisition_fit_score":  int(round(scored.get("total_score", 0))),
        "important_notes":        compose_notes(row, category, scored),
    }


# ── Excel export ──────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT     = Font(name="Arial", size=10)
SCORE_HIGH    = PatternFill("solid", fgColor="C6EFCE")
SCORE_MED     = PatternFill("solid", fgColor="FFEB9C")
SCORE_LOW     = PatternFill("solid", fgColor="FFC7CE")
THIN_BORDER   = Border(bottom=Side(style="thin", color="D9D9D9"))

COL_WIDTHS = {
    "business_name": 32, "address": 30, "city": 14, "postal_code": 12,
    "website": 28, "phone": 18, "owner_name": 22, "owner_confidence": 14,
    "owner_source": 35, "industry": 18, "category_standardized": 28,
    "employee_range_estimate": 16, "revenue_range_estimate": 18,
    "sde_range_estimate": 18, "age_range_estimate": 16,
    "acquisition_fit_score": 14, "important_notes": 58,
}

STANDARD_FIELDS = [
    "business_name", "address", "city", "postal_code", "website", "phone",
    "owner_name", "owner_confidence", "owner_source", "industry",
    "category_standardized", "employee_range_estimate", "revenue_range_estimate",
    "sde_range_estimate", "age_range_estimate", "acquisition_fit_score", "important_notes",
]


def export_excel(df, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "New Non-Trades Leads"
    ws.freeze_panes = "B2"

    for col_idx, header in enumerate(STANDARD_FIELDS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, header in enumerate(STANDARD_FIELDS, 1):
            value = row.get(header, "")
            if pd.isna(value):
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center",
                                       wrap_text=(header == "important_notes"))

            if header == "acquisition_fit_score" and value != "":
                try:
                    score = int(float(value))
                    cell.alignment = Alignment(horizontal="center")
                    cell.fill = SCORE_HIGH if score >= 70 else (SCORE_MED if score >= 50 else SCORE_LOW)
                except (ValueError, TypeError):
                    pass

            if header == "owner_confidence":
                cell.alignment = Alignment(horizontal="center")

    for col_idx, header in enumerate(STANDARD_FIELDS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(header, 15)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(STANDARD_FIELDS))}{len(df)+1}"
    ws.row_dimensions[1].height = 30

    wb.save(path)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print(" STEP 10: NEW NON-TRADES BATCH (10 Fresh Leads)")
    print("=" * 60)

    if GOOGLE_PLACES_API_KEY == "YOUR_KEY_HERE":
        print("[ERROR] GOOGLE_PLACES_API_KEY not set in .env")
        sys.exit(1)

    # Load existing place_ids to exclude
    existing_ids = set()
    if os.path.exists(CHECKPOINT_RAW):
        existing_df = pd.read_csv(CHECKPOINT_RAW, usecols=["google_place_id"])
        existing_ids = set(existing_df["google_place_id"].dropna().astype(str))
        print(f"  Excluding {len(existing_ids)} existing place_ids from {CHECKPOINT_RAW}")

    # Fetch new candidates
    print(f"\n  Fetching {len(NEW_BATCH_KEYWORDS)} keyword searches...")
    all_records = []
    seen_ids = set(existing_ids)
    live_calls = 0

    for kw in NEW_BATCH_KEYWORDS:
        records, was_live = fetch_keyword(kw, seen_ids)
        if was_live:
            live_calls += 1
        added = len(records)
        all_records.extend(records)
        if added:
            print(f"    [{kw:35s}] +{added:>3} new  (total so far: {len(all_records)})")
        time.sleep(GOOGLE_DELAY_SECONDS)

    print(f"\n  Total raw candidates fetched: {len(all_records)}")
    print(f"  Live API calls: {live_calls}  (est. cost: ${live_calls * 0.032:.2f})")

    if not all_records:
        print("[WARN] No new candidates found. Check keywords or API key.")
        return

    df = pd.DataFrame(all_records)

    # Filter
    print("\n  Applying filters...")
    df = apply_filters(df)
    print(f"  Candidates after filtering: {len(df)}")

    if df.empty:
        print("[WARN] No candidates passed filters.")
        return

    # Fill employee estimates
    for idx, row in df.iterrows():
        emp = row.get("num_employees")
        if emp is None or pd.isna(emp):
            result = estimate_employee_range(row)
            if result:
                df.at[idx, "num_employees"] = result["employee_midpoint"]

        # Fill estimated_years from review count
        reviews = row.get("review_count", 0)
        age_result = estimate_business_age(reviews)
        if age_result:
            df.at[idx, "estimated_years"] = age_result["estimated_years"]

    # Score
    print("\n  Scoring candidates...")
    score_cols = []
    for _, row in df.iterrows():
        industry = get_industry_label(row)
        scored = score_row(row, industry)
        score_cols.append(scored)

    scores_df = pd.DataFrame(score_cols)
    df = df.reset_index(drop=True)
    df = pd.concat([df, scores_df], axis=1)
    df = df.sort_values("total_score", ascending=False).reset_index(drop=True)

    print(f"\n  Top 20 candidates by score:")
    for _, row in df.head(20).iterrows():
        print(f"    {row['total_score']:>5.1f}  {row['company_name'][:50]}"
              f"  [{row.get('industry_label', '')}]")

    # Pick top 10
    top10 = df.head(10).copy()

    # Transform to Hamilton standard format
    leads = []
    for _, row in top10.iterrows():
        scored = {
            "industry_label": row.get("industry_label", "general"),
            "estimated_revenue_low": row.get("estimated_revenue_low", 0),
            "estimated_revenue_high": row.get("estimated_revenue_high", 0),
            "estimated_sde_low": row.get("estimated_sde_low", 0),
            "estimated_sde_high": row.get("estimated_sde_high", 0),
            "revenue_confidence": row.get("revenue_confidence", 20),
            "total_score": row.get("total_score", 0),
        }
        leads.append(to_standard_row(row, scored))

    output_df = pd.DataFrame(leads, columns=STANDARD_FIELDS)

    # Print summary
    print(f"\n  Final 10 leads:")
    for i, (_, row) in enumerate(output_df.iterrows(), 1):
        print(f"    {i:>2}. [{row['acquisition_fit_score']:>3}]  "
              f"{row['business_name'][:45]:<45}  {row['category_standardized']}")

    # Export Excel
    os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)
    export_excel(output_df, OUTPUT_EXCEL)
    print(f"\n  [OUTPUT] Saved -> {OUTPUT_EXCEL}")
    print(f"  [OUTPUT] {len(output_df)} leads, 17-field Hamilton standard format")


if __name__ == "__main__":
    main()
